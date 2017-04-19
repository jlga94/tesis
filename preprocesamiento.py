from nltk.corpus import stopwords
from stop_words import get_stop_words
import csv
import operator
import codecs
import string
import math
import unicodedata
import treetaggerwrapper
import openpyxl
import re
import numpy
import warnings
from nltk.stem.snowball import SpanishStemmer
from langdetect import detect
from nltk.tokenize import sent_tokenize
from gensim.models import Phrases,Word2Vec


class TextProcessor:
    lemmatizer=None
    stopEnglish=None
    stopSpanish=None
    spanishStemmer=None

    def __init__(self):
        self.lemmatizer = treetaggerwrapper.TreeTagger(TAGLANG='es')
        self.lemmaTypeWords = {'adv','ord','ppc','ppx','ppo','qu','rel','ummx','vcliger','vcliinf','vclifin','veadj','vefin','veger','veinf',
        'vhadj','vhfin','vhger','vhinf','vladj','vlfin','vlger','vlinf','vmadj','vmfin','vmger','vminf','vsadj','vsfin','vsger','vsinf'}

        stopwords_en_nltk = set(stopwords.words('english'))
        stopwords_en_sw = set(get_stop_words('english'))
        self.stopEnglish = set(stopwords_en_nltk.union(stopwords_en_sw))
        
        stopwords_sp_nltk = set(stopwords.words('spanish'))
        stopwords_sp_sw = set(get_stop_words('spanish'))
        self.stopSpanish = set(stopwords_sp_nltk.union(stopwords_sp_sw))
        additionalStopwordsSpanish = {'u','y/o','año','años','alto','ser','etc','respecto','hacer','tal','dentro','mes','meses','tener','experiencia','trabajo','bueno','afín','afine',
        'nivel','pequeño','haber','menos','menor','deseable','incluir','pues','parte','manera','según','lunes','martes','miércoles','jueves','viernes','sábado','domingo','lugar','fondo',
        'enero','febrero','marzo','abril','mayo','junio','julio','agosto','setiembre','octubre','noviembre','diciembre','asi','así','través','uno','uso','casa','mismo','mediante','gran',
        'grande','hacia','conforme','número','siguiente','link','cuatro','tres','cinco','sitio','lista','anual','mensual','trimestral','bimestral','semestral','semanal','diario','día'}
        self.stopSpanish = self.stopSpanish.union(additionalStopwordsSpanish)
        
        self.spanishStemmer=SpanishStemmer()

        
    def _remove_numbers(self, text):
        "Elimina los números del texto"

        return ''.join([letter for letter in text if not letter.isdigit()])

    def _remove_punctuation(self, text):
        "Elimina los signos de puntuacion del texto"

        regex = re.compile('[%s]' % re.escape(string.punctuation))
        return regex.sub(' ', text)

    def _separateFirstLetterFromPunctuation(self,text):
    	newText=''
    	previousLetter='1'
    	for letter in text:
    		if(previousLetter in string.punctuation and letter.isalpha()):
    			newText+=' '   			
    		newText+=letter
    		previousLetter=letter
    	return newText

    def removeStopwordsInList(self,text):
    	return [word for word in text.split() if word not in self.stopEnglish and word not in self.stopSpanish]

    def _retireRareLetters(self,text):
    	return ''.join([letter for letter in text if letter.isalpha() or letter in string.punctuation or letter==' ' or letter=='\n'])

    def _retireUrlsWWW(self,text):
    	return re.sub(r'\s*(?:https?://)?www\.\S*\.[A-Za-z]{2,5}\s*', '', text).strip()

    def _retireUrlsHTTP(self,text):
    	return re.sub(r'\w+:\/{2}[\d\w-]+(\.[\d\w-]+)*(?:(?:\/[^\s/]*))*', '', text).strip()

    def firstpreprocessText(self,text):
    	text=text.lower()
    	text=self._retireRareLetters(text)
    	text=self._retireUrlsWWW(text)
    	text=self._retireUrlsHTTP(text)
    	return text

    def preprocessText(self,text):
    	text=self._remove_punctuation(text)
    	text=self._remove_numbers(text)
    	return text

    def lematizeText(self,text):
    	newText = ""
    	lemmaElement = 2

    	for sentence in sent_tokenize(text,language='spanish'): #Para analizar frases y mejorar el desempeno del lemmatizador
    		textWithSeparatePunctuation=self._separateFirstLetterFromPunctuation(sentence)
    		lemmaResultText = self.lemmatizer.tag_text(textWithSeparatePunctuation)# Return [[word,type of word, lemma]]
    		#print(lemmaResultText)
    		for wordLemmaResult in lemmaResultText:
    			#print(word)
    			word = wordLemmaResult.split('\t')[lemmaElement].lower().strip()    			
    			if word not in self.stopEnglish and word not in self.stopSpanish: #Se retiran stopwords
    				if not(len(word)==1 and not(word in string.punctuation)):
    					newText += ' ' + word
    	
    	return newText.strip()

    def lematizeText_Filter_TypeWord(self,text):
    	newText = ""
    	lemmaTypeWord = 1
    	lemmaElement = 2

    	for sentence in sent_tokenize(text,language='spanish'): #Para analizar frases y mejorar el desempeno del lemmatizador
    		textWithSeparatePunctuation=self._separateFirstLetterFromPunctuation(sentence)
    		lemmaResultText = self.lemmatizer.tag_text(textWithSeparatePunctuation)# Return [[word,type of word, lemma]]
    		#print(lemmaResultText)
    		for wordLemmaResult in lemmaResultText:
    			#print(word)
    			typeWord = wordLemmaResult.split('\t')[lemmaTypeWord].lower().strip()
    			if typeWord not in self.lemmaTypeWords and typeWord[0]!='v':
	    			word = wordLemmaResult.split('\t')[lemmaElement].lower().strip()
	    			if word not in self.stopEnglish and word not in self.stopSpanish: #Se retiran stopwords
	    				if not(len(word)==1 and not(word in string.punctuation)):
	    					newText += ' ' + word
    	
    	return newText.strip()

    def stemText_Filter_TypeWord(self,text):
    	newText = ""
    	lemmaTypeWord = 1
    	lemmaOriginalWord = 0

    	for sentence in sent_tokenize(text,language='spanish'): #Para analizar frases y mejorar el desempeno
    		textWithSeparatePunctuation=self._separateFirstLetterFromPunctuation(sentence)
    		lemmaResultText = self.lemmatizer.tag_text(textWithSeparatePunctuation)# Return [[word,type of word, lemma]]
    		for wordLemmaResult in lemmaResultText:
    			#print(word)
    			typeWord = wordLemmaResult.split('\t')[lemmaTypeWord].lower().strip()
    			if typeWord not in self.lemmaTypeWords and typeWord[0]!='v':
	    			word = wordLemmaResult.split('\t')[lemmaOriginalWord].lower().strip()
	    			word = word.replace("\ufeff", "")
	    			if word not in self.stopEnglish and word not in self.stopSpanish: #Se retiran stopwords
	    				if not(len(word)==1 and not(word in string.punctuation)):
	    					wordStemmed = self.spanishStemmer.stem(word)
    						newText += " " + wordStemmed

    	return newText.strip()

    def stemText(self,text):
    	text=self._separateFirstLetterFromPunctuation(text)
    	newText = ""
    	for word in text.split():
    		if word not in self.stopEnglish and word not in self.stopSpanish:
    			word = word.replace("\ufeff", "")
    			if not(len(word)==1 and not(word in string.punctuation)):
    				wordStemmed = self.spanishStemmer.stem(word)
    				newText += " " + wordStemmed
    	return newText.strip()

class Input_Output:
    procesadorTextos=None

    def __init__(self):
        self.procesadorTextos=TextProcessor()
        self.secciones={}
        self.nombColumnas=[]

    def limpiarDataset_Steeming(self, dataset):
    	newDataset = []
    	for ofertaText in dataset:
    		ofertaText = self.procesadorTextos.preprocessText(ofertaText)
    		ofertaText = self.procesadorTextos.stemText(ofertaText)
    		listaPalabras = ofertaText.strip().split()
    		newDataset.append(listaPalabras)
    	print('Se termino de limpiarlo completamente con Steeming.')
    	return newDataset

    def limpiarDataset_Lemmatizacion(self, dataset):
    	newDataset = []
    	for oferta in dataset:
    		cadenaLemmatizada = self.procesadorTextos.lematizeText(oferta.strip())
    		cadenaPreProcesadaLemmatizada = self.procesadorTextos.preprocessText(cadenaLemmatizada)
    		listaPalabras = cadenaPreProcesadaLemmatizada.split()
    		newDataset.append(listaPalabras)
    	print('Se termino de limpiarlo completamente con Lemmatizacion.')
    	return newDataset

    def limpiarDataset_Steeming_Bigram(self, dataset,thresholds):
    	bigrams = []
    	for threshold in thresholds:
    		bigrams.append(Phrases(threshold=threshold))
    	newDataset = []
    	numOferta = 1
    	for ofertaText in dataset:
    		ofertaText = self.procesadorTextos.preprocessText(ofertaText)
    		#ofertaText = self.procesadorTextos.stemText(ofertaText)
    		ofertaText = self.procesadorTextos.stemText_Filter_TypeWord(ofertaText)
    		listaPalabras = self.procesadorTextos.removeStopwordsInList(ofertaText.strip())
    		if len(listaPalabras)>0:
    			for bigram in bigrams:
    				bigram.add_vocab([listaPalabras])
    			newDataset.append(listaPalabras)
    		else:
    			print('Stem: No se tomo en cuenta la oferta '+ str(numOferta))
    		numOferta+=1
    	print('Se termino de limpiarlo completamente con Steeming-Bigram.')
    	return newDataset,bigrams

    def limpiarDataset_Lemmatizacion_Bigram(self, dataset,thresholds):
    	bigrams = []
    	for threshold in thresholds:
    		bigrams.append(Phrases(threshold=threshold)) 
    	
    	newDataset = []
    	numOferta = 1
    	for oferta in dataset:  		
    		cadenaLemmatizada = self.procesadorTextos.lematizeText_Filter_TypeWord(oferta.strip())
    		#cadenaLemmatizada = self.procesadorTextos.lematizeText(oferta.strip())
    		cadenaPreProcesadaLemmatizada = self.procesadorTextos.preprocessText(cadenaLemmatizada)
    		listaPalabras = self.procesadorTextos.removeStopwordsInList(cadenaPreProcesadaLemmatizada)
    		if len(listaPalabras)>0:
    			for bigram in bigrams:
    				bigram.add_vocab([listaPalabras])
    			newDataset.append(listaPalabras)
    		else:
    			print('Lemma: No se tomo en cuenta la oferta '+ str(numOferta))
    		numOferta+=1
    	print('Se termino de limpiarlo completamente con Lemmatizacion-Bigram.')
    	return newDataset,bigrams

    def leerSecciones(self, filename):
    	dataset = []
    	wb = openpyxl.load_workbook(filename)
    	sheets = wb.get_sheet_names()
    	sheetAviso = wb.get_sheet_by_name(sheets[0])
    	maxColumnas = sheetAviso.max_column + 1
    	primeraFila = 1
    	for iColumna in range(1,maxColumnas):
    		seccion = str(sheetAviso.cell(row=primeraFila, column=iColumna).value)
    		self.nombColumnas.append(seccion)
    		self.secciones[seccion]=iColumna
    
    def buscarNombre_Descripcion_Conocimientos_Ofertas():
    	columnasABuscar=[]
    	columnasABuscar.append(self.secciones['Job: Job Title'])
    	columnasABuscar.append(self.secciones['Job: Description'])
    	columnasABuscar.append(self.secciones['Job: Qualifications'])
    	return columnasABuscar

    def obtenerDatasetAClasificar_Completo(self, filename):
        "Se lee un archivo Excel con las ofertas a clasificar"

        dataset = []
        wb = openpyxl.load_workbook(filename)
        sheets = wb.get_sheet_names()
        sheetAviso = wb.get_sheet_by_name(sheets[0])
        maxFilas = sheetAviso.max_row + 1

        for num_oferta in range(1, maxFilas):
            text = str(sheetAviso.cell(row=num_oferta, column=1).value)
            dataset.append(text)
        
        datasetLemmatizacion = self.limpiarDataset_Lemmatizacion(dataset)
        #datasetSteeming = self.limpiarDataset_Steeming(dataset)

        #return datasetSteeming
        return datasetLemmatizacion
        #return datasetLemmatizacion,datasetSteeming

    def obtenerDatasetAClasificar_Completo_Bigram(self, filename,thresholds_Lemma,thresholds_Stem):
        "Se lee un archivo Excel con las ofertas a clasificar"

        dataset = []
        wb = openpyxl.load_workbook(filename)
        sheets = wb.get_sheet_names()
        sheetAviso = wb.get_sheet_by_name(sheets[0])
        maxFilas = sheetAviso.max_row + 1
        infoColumn = 1

        for num_oferta in range(1, maxFilas):
            text = str(sheetAviso.cell(row=num_oferta, column=infoColumn).value)
            dataset.append(text)

        datasetLemmatizacion,bigramsLemmatizacion = self.limpiarDataset_Lemmatizacion_Bigram(dataset,thresholds_Lemma)
        datasetSteeming,bigramsSteeming = self.limpiarDataset_Steeming_Bigram(dataset,thresholds_Stem)

        #return datasetSteeming
        return datasetLemmatizacion,bigramsLemmatizacion,datasetSteeming,bigramsSteeming
        #return datasetLemmatizacion,datasetSteeming

    def obtenerDatasetAClasificar_Dividido(self, filename):
    	dataset_Title_Descp = []
    	dataset_Qualif = []
    	wb = openpyxl.load_workbook(filename)
    	sheets = wb.get_sheet_names()
    	sheetAviso = wb.get_sheet_by_name(sheets[0])
    	maxFilas = sheetAviso.max_row + 1
    	for num_oferta in range(1, maxFilas):
    		text_Title_Descp = str(sheetAviso.cell(row=num_oferta, column=1).value)
    		dataset_Title_Descp.append(text_Title_Descp)
    		
    		text_Qualif = str(sheetAviso.cell(row=num_oferta, column=2).value)
    		if (text_Qualif!=''):
    			dataset_Qualif.append(text_Qualif)

    	datasetLemmatizacion_Title_Descp = self.limpiarDataset_Lemmatizacion(dataset_Title_Descp)
    	datasetLemmatizacion_Qualif = self.limpiarDataset_Lemmatizacion(dataset_Qualif)

    	datasetSteeming_Title_Descp = self.limpiarDataset_Steeming(dataset_Title_Descp)
    	datasetSteeming_Qualif = self.limpiarDataset_Steeming(dataset_Qualif)

    	return datasetLemmatizacion_Title_Descp,datasetLemmatizacion_Qualif,datasetSteeming_Title_Descp,datasetSteeming_Qualif

    def contarIdiomasEnOfertas(self, filename):
    	dataset = []
    	wb = openpyxl.load_workbook(filename)
    	sheets = wb.get_sheet_names()
    	sheetAviso = wb.get_sheet_by_name(sheets[0])
    	maxFilas = sheetAviso.max_row + 1
    	maxColumnas = sheetAviso.max_column + 1

    	self.leerSecciones(filename)
    	columnasABuscar=[]
    	columnasABuscar.append(self.secciones['Job: Job Title'])
    	columnasABuscar.append(self.secciones['Job: Description'])
    	columnasABuscar.append(self.secciones['Job: Qualifications'])

    	ofertasRaras={}
    	idiomas={}
    	for num_oferta in range(2, maxFilas):
    		completeText = ''
    		for nColumna in columnasABuscar:
    			text = str(sheetAviso.cell(row=num_oferta, column=nColumna).value)
    			#completeText+=self.procesadorTextos.firstpreprocessText(text)
    			completeText+=' '+text.lower()
    		idioma = detect(completeText.strip())
    		if(idioma not in idiomas.keys()):
    			idiomas[idioma]=1
    		else:
    			idiomas[idioma]+=1
    		if(idioma!='es' and idioma!='en'):
    			if(idioma not in ofertasRaras.keys()):
    				ofertasRaras[idioma]=[completeText]
    			else:
    				ofertasRaras[idioma].append(completeText)
    	return idiomas,ofertasRaras
            


    def primeraLimpiezaADatasetAClasificar_Completo(self, filename):
    	wb = openpyxl.load_workbook(filename)
    	sheets = wb.get_sheet_names()
    	sheetAviso = wb.get_sheet_by_name(sheets[0])
    	maxFilas = sheetAviso.max_row + 1
    	maxColumnas = sheetAviso.max_column + 1

    	newExcel = openpyxl.Workbook()
    	newSheet = newExcel.active

    	self.leerSecciones(filename)
    	columnasABuscar=[]
    	columnasABuscar.append(self.secciones['Job: Job Title'])
    	columnasABuscar.append(self.secciones['Job: Description'])
    	columnasABuscar.append(self.secciones['Job: Qualifications'])

    	numColNewExcel_Text = 1
    	numColNewExcel_Year = 2
    	numColNewExcel_Month = 3

    	numRowNewExcel=1
    	for num_oferta in range(2, maxFilas):
    		completeText = ''
    		for nColumna in columnasABuscar:
    			text = str(sheetAviso.cell(row=num_oferta, column=nColumna).value)
    			completeText+=' ' + text.lower()
    		idioma = detect(completeText.strip())
    		if(idioma=='es'):
    			text = self.procesadorTextos.firstpreprocessText(completeText)
    			newSheet.cell(row=numRowNewExcel, column=numColNewExcel_Text).value = text

    			date = str(sheetAviso.cell(row=num_oferta, column=self.secciones['Job: Posting Date']).value)
    			print(date)
    			ddmmyyyy = date.split('/')

    			newSheet.cell(row=numRowNewExcel, column=numColNewExcel_Year).value = int(ddmmyyyy[2])
    			newSheet.cell(row=numRowNewExcel, column=numColNewExcel_Month).value = int(ddmmyyyy[1])

    			numRowNewExcel+=1

    	filenameWithoutExtension = filename.split('.')[0]
    	newExcel.save(filenameWithoutExtension + '_PrimerPreProcesamiento_Completo.xlsx')

    def primeraLimpiezaADatasetAClasificar_Dividido(self, filename):
    	wb = openpyxl.load_workbook(filename)
    	sheets = wb.get_sheet_names()
    	sheetAviso = wb.get_sheet_by_name(sheets[0])
    	maxFilas = sheetAviso.max_row + 1
    	maxColumnas = sheetAviso.max_column + 1

    	newExcel = openpyxl.Workbook()
    	newSheet = newExcel.active

    	self.leerSecciones(filename)
    	columnasABuscar=[]
    	columnasABuscar.append(self.secciones['Job: Job Title'])
    	columnasABuscar.append(self.secciones['Job: Description'])
    	#columnasABuscar.append(self.secciones['Job: Qualifications'])

    	nColQualifications = self.secciones['Job: Qualifications']

    	numColNewExcel_Title_Descp=1
    	numColNewExcel_Qualifications=2
    	numColNewExcel_Year=3
    	numColNewExcel_Month = 4

    	numRowNewExcel=1
    	for num_oferta in range(2, maxFilas):
    		title_descp_Text = ''
    		qualif_Text = str(sheetAviso.cell(row=num_oferta, column=nColQualifications).value).lower().strip()

    		for nColumna in columnasABuscar:
    			text = str(sheetAviso.cell(row=num_oferta, column=nColumna).value)
    			title_descp_Text+=' ' + text.lower()
    		title_descp_Text.strip()

    		idioma = detect(title_descp_Text+' '+qualif_Text) # Debe hacerse al texto completo
    		if(idioma=='es'):
    			title_descp_Text = self.procesadorTextos.firstpreprocessText(title_descp_Text)
    			qualif_Text = self.procesadorTextos.firstpreprocessText(qualif_Text)
    			
    			newSheet.cell(row=numRowNewExcel, column=numColNewExcel_Title_Descp).value = title_descp_Text
    			newSheet.cell(row=numRowNewExcel, column=numColNewExcel_Qualifications).value = qualif_Text
    			
    			date = str(sheetAviso.cell(row=num_oferta, column=self.secciones['Job: Posting Date']).value)
    			print(date)
    			ddmmyyyy = date.split('/')

    			newSheet.cell(row=numRowNewExcel, column=numColNewExcel_Year).value = int(ddmmyyyy[2])
    			newSheet.cell(row=numRowNewExcel, column=numColNewExcel_Month).value = int(ddmmyyyy[1])
    			
    			numRowNewExcel+=1

    	filenameWithoutExtension = filename.split('.')[0]
    	newExcel.save(filenameWithoutExtension + '_PrimerPreProcesamiento_Dividido.xlsx')

    def primeraLimpiezaADatasetAClasificar_D_Q(self, filename):
    	wb = openpyxl.load_workbook(filename)
    	sheets = wb.get_sheet_names()
    	sheetAviso = wb.get_sheet_by_name(sheets[0])
    	maxFilas = sheetAviso.max_row + 1
    	maxColumnas = sheetAviso.max_column + 1

    	newExcel = openpyxl.Workbook()
    	newSheet = newExcel.active

    	self.leerSecciones(filename)
    	columnasABuscar=[]
    	#columnasABuscar.append(self.secciones['Job: Job Title'])
    	columnasABuscar.append(self.secciones['Job: Description'])
    	columnasABuscar.append(self.secciones['Job: Qualifications'])

    	numColNewExcel_Text = 1
    	numColNewExcel_Year = 2
    	numColNewExcel_Month = 3

    	numRowNewExcel=1
    	for num_oferta in range(2, maxFilas):
    		completeText = ''
    		for nColumna in columnasABuscar:
    			text = str(sheetAviso.cell(row=num_oferta, column=nColumna).value)
    			completeText+=' ' + text.lower()
    		idioma = detect(completeText.strip())
    		if(idioma=='es'):
    			text = self.procesadorTextos.firstpreprocessText(completeText)
    			newSheet.cell(row=numRowNewExcel, column=numColNewExcel_Text).value = text

    			date = str(sheetAviso.cell(row=num_oferta, column=self.secciones['Job: Posting Date']).value)
    			print(date)
    			ddmmyyyy = date.split('/')

    			newSheet.cell(row=numRowNewExcel, column=numColNewExcel_Year).value = int(ddmmyyyy[2])
    			newSheet.cell(row=numRowNewExcel, column=numColNewExcel_Month).value = int(ddmmyyyy[1])

    			numRowNewExcel+=1

    	filenameWithoutExtension = filename.split('.')[0]
    	newExcel.save(filenameWithoutExtension + '_PrimerPreProcesamiento_D_Q.xlsx')

