from gensim import corpora,models
import gensim
from preprocesamiento import Input_Output
from collections import Counter
from gensim.models import Phrases,Word2Vec,CoherenceModel
import openpyxl, pickle, operator

coherenceTypes = ['u_mass','c_v','c_uci','c_npmi']

def writeLDAFile(numArch,tipo,listTopics,dictionary,corpus,texts):	
	coherencePerKTopic = {}
	for i in range(len(listTopics)):
		print('Empezando '+ tipo +' con: '+str(listTopics[i]))
		ldamodel = gensim.models.ldamodel.LdaModel(corpus, num_topics=listTopics[i],id2word = dictionary, passes=20)
		print('Se creo objeto')
		
		F = open(numArch + '_'+ tipo +'_'+str(listTopics[i])+'.txt','w')

		coherenceTypesValues = []
		for iCoherenceType in coherenceTypes:
			cm_coherence_Model = CoherenceModel(model=ldamodel,corpus=corpus,texts=texts,coherence=iCoherenceType).get_coherence()
			coherenceTypesValues.append(cm_coherence_Model)
			F.write('Coherencia '+ str(iCoherenceType) +' de ' + numArch + '_'+ tipo +'_'+str(listTopics[i]) + ': ' +str(cm_coherence_Model)+'\n')
		
		coherencePerKTopic[listTopics[i]] = coherenceTypesValues

		#F.write(str(ldamodel.print_topics(num_topics=listTopics[i],num_words=15))+'\n')
		for iTopic in range(listTopics[i]):
			F.write('Topic '+str(iTopic)+': '+str(ldamodel.print_topic(topicno=iTopic,topn=15))+'\n')
		F.close()
		print('Terminando '+ tipo +' con: '+str(listTopics[i]))

	return coherencePerKTopic

def writeExcelCoherencePerTopic(carrera,tipo,thresholds,listTopics,coherencePerThreshold):
	newExcel = openpyxl.Workbook()
	actualSheet = newExcel.active
	ThresholdColumn = 1
	K_Column = 2
	coherenceColumns = [3,4,5,6]

	actualSheet.title = carrera + '_' + tipo
	actualSheet.cell(row=1, column=ThresholdColumn).value = 'Threshold'
	actualSheet.cell(row=1, column=K_Column).value = 'K_Topic'
	for coherenceIndex in range(len(coherenceColumns)):
		actualSheet.cell(row=1, column=coherenceColumns[coherenceIndex]).value = coherenceTypes[coherenceIndex]

	actualRow = 2
	for threshold in thresholds:
		actualSheet.cell(row=actualRow, column=ThresholdColumn).value = threshold
		for topicIndex in range(len(listTopics)):
			actualSheet.cell(row=actualRow, column=K_Column).value = listTopics[topicIndex]
			for coherenceIndex in range(len(coherenceTypes)):
				actualSheet.cell(row=actualRow, column=coherenceColumns[coherenceIndex]).value = coherencePerThreshold[threshold][listTopics[topicIndex]][coherenceIndex]
			actualRow += 1

	newExcel.save(carrera + '_' + tipo + '_Resultados_Coherencia.xlsx')

def LDA_GENSIM_COMPLETO():
	inputOutput = Input_Output()
	carrera = 'Contabilidad'

	datasetLemmatizacion,datasetSteeming = inputOutput.obtenerDatasetAClasificar_Completo('Avisos_'+carrera+'_2011-2016_PrimerPreProcesamiento_Completo.xlsx')

	dictionaryLemma = corpora.Dictionary(datasetLemmatizacion)
	corpusLemma = [dictionaryLemma.doc2bow(text) for text in datasetLemmatizacion]

	dictionarySteeming = corpora.Dictionary(datasetSteeming)
	corpusSteeming = [dictionarySteeming.doc2bow(text) for text in datasetSteeming]

	listTopics= [10,15,20,25]
	listNumWords = []

	#ldamodelLemma = gensim.models.ldamodel.LdaModel(corpusLemma, num_topics=20,id2word = dictionaryLemma, passes=20)

	writeLDAFile(carrera,'Lemmatizacion',listTopics,dictionaryLemma,corpusLemma)
	writeLDAFile(carrera,'Steeming',listTopics,dictionarySteeming,corpusSteeming)


def LDA_GENSIM_DIVIDIDO():
	inputOutput = Input_Output()
	carrera = 'Industrial'

	datasetLemmatizacion_Title_Descp,datasetLemmatizacion_Qualif,datasetSteeming_Title_Descp,datasetSteeming_Qualif = inputOutput.obtenerDatasetAClasificar_Dividido('Avisos_'+carrera+'_2011-2016_PrimerPreProcesamiento_Dividido.xlsx')

	dictionaryLemma_T_D = corpora.Dictionary(datasetLemmatizacion_Title_Descp)
	corpusLemma_T_D = [dictionaryLemma_T_D.doc2bow(text) for text in datasetLemmatizacion_Title_Descp]

	dictionaryLemma_Q = corpora.Dictionary(datasetLemmatizacion_Qualif)
	corpusLemma_Q = [dictionaryLemma_Q.doc2bow(text) for text in datasetLemmatizacion_Qualif]

	dictionarySteeming_T_D = corpora.Dictionary(datasetSteeming_Title_Descp)
	corpusSteeming_T_D = [dictionarySteeming_T_D.doc2bow(text) for text in datasetSteeming_Title_Descp]

	dictionarySteeming_Q = corpora.Dictionary(datasetSteeming_Qualif)
	corpusSteeming_Q = [dictionarySteeming_Q.doc2bow(text) for text in datasetSteeming_Qualif]

	listTopics= [10,15,20,25]
	
	#ldamodelLemma = gensim.models.ldamodel.LdaModel(corpusLemma, num_topics=20,id2word = dictionaryLemma, passes=20)

	writeLDAFile(carrera,'Lemmatizacion_T_D',listTopics,dictionaryLemma_T_D,corpusLemma_T_D)
	writeLDAFile(carrera,'Lemmatizacion_Q',listTopics,dictionaryLemma_Q,corpusLemma_Q)
	writeLDAFile(carrera,'Steeming_T_D',listTopics,dictionarySteeming_T_D,corpusSteeming_T_D)
	writeLDAFile(carrera,'Steeming_Q',listTopics,dictionarySteeming_Q,corpusSteeming_Q)

def replaceBigrams(dataset,bigram_model):
	newDataset=[]
	for words_oferta in dataset:
		N_words = len(words_oferta)
		i=0
		while(i < N_words-1):
			possible_bigram = words_oferta[i] + '_' +words_oferta[i+1]
			if (possible_bigram in bigram_model.wv.vocab.keys()):
				words_oferta[i] = possible_bigram # Se reemplaza la palabra del indice i por el bigrama
				words_oferta.pop(i+1)
				N_words-=1
			i+=1
		newDataset.append(words_oferta)
	return newDataset

def replaceBigrams_v2(dataset,bigram_model):
	newDataset=[]
	for words_oferta in dataset:
		N_words = len(words_oferta)
		i=0
		while(i < N_words-1):
			possible_bigram = words_oferta[i] + '_' +words_oferta[i+1]
			if (possible_bigram in bigram_model.wv.vocab.keys()):
				words_oferta.insert(i,possible_bigram) #se agrega el bigrama como indice sin borrar los unigramas
				i+=1
				N_words+=1
			i+=1
		newDataset.append(words_oferta)
	return newDataset

def LDA_GENSIM_COMPLETO_BIGRAM():
	inputOutput = Input_Output()
	carrera = 'Informatica'
	thresholds = [15,16,17,18,19,20,21,22,23,24,25]

	datasetLemmatizacion,bigramsLemmatizacion,datasetSteeming,bigramsSteeming = inputOutput.obtenerDatasetAClasificar_Completo_Bigram('Avisos_'+carrera+'_2011-2016_PrimerPreProcesamiento_Completo.xlsx',thresholds)

	bigram_models_Lemmatizacion = []
	bigram_models_Steeming = []

	print('Comenzo bigram Lemma')
	for bigramLemmatizacion in bigramsLemmatizacion:
		
		bigram_model_Lemmatizacion = Word2Vec(bigramLemmatizacion[datasetLemmatizacion],size=100)
		bigram_models_Lemmatizacion.append(bigram_model_Lemmatizacion)
	
	print('Comenzo bigram Steeming')
	for bigramSteeming in bigramsSteeming:		
		bigram_model_Steeming = Word2Vec(bigramSteeming[datasetSteeming],size=100)
		bigram_models_Steeming.append(bigram_model_Steeming)

	datasetsLemmatizacionBigram = []
	datasetsSteemingBigram = []

	print('Comenzo bigram Lemma Reemplazo')
	for bigram_model_Lemmatizacion in bigram_models_Lemmatizacion:
		datasetLemmatizacionBigram = replaceBigrams(datasetLemmatizacion,bigram_model_Lemmatizacion)
		datasetsLemmatizacionBigram.append(datasetLemmatizacionBigram)
	
	print('Comenzo bigram Steeming Reemplazo')
	for bigram_model_Steeming in bigram_models_Steeming:
		datasetSteemingBigram = replaceBigrams(datasetSteeming,bigram_model_Steeming)
		datasetsSteemingBigram.append(datasetSteemingBigram)

	dictionariesLemmaPerThreshold = []
	corpusLemmaPerThreshold = []
	print('Comenzo a obtener diccionarios y corpus para Gensim de Lemmatizacion')
	for datasetLemmatizacionBigram in datasetsLemmatizacionBigram:
		dictionaryLemma = corpora.Dictionary(datasetLemmatizacionBigram)
		corpusLemma = [dictionaryLemma.doc2bow(text) for text in datasetLemmatizacionBigram]
		dictionariesLemmaPerThreshold.append(dictionaryLemma)
		corpusLemmaPerThreshold.append(corpusLemma)

	dictionariesSteemingPerThreshold = []
	corpusSteemingPerThreshold = []
	print('Comenzo a obtener diccionarios y corpus para Gensim de Steeming')
	for datasetSteemingBigram in datasetsSteemingBigram:
		dictionarySteeming = corpora.Dictionary(datasetSteemingBigram)
		corpusSteeming = [dictionarySteeming.doc2bow(text) for text in datasetSteemingBigram]
		dictionariesSteemingPerThreshold.append(dictionarySteeming)
		corpusSteemingPerThreshold.append(corpusSteeming)


	listTopics= [4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23]

	coherencePerThresholdLemma = {}
	coherencePerThresholdSteeming = {}


	for thresholdIndex in range(len(thresholds)):
		coherencePerKTopic = writeLDAFile(carrera,'Lemmatizacion_Bigram_Th_'+str(thresholds[thresholdIndex]),listTopics,dictionariesLemmaPerThreshold[thresholdIndex],corpusLemmaPerThreshold[thresholdIndex],datasetsLemmatizacionBigram[thresholdIndex])
		coherencePerThresholdLemma[thresholds[thresholdIndex]] = coherencePerKTopic

		coherencePerKTopic = writeLDAFile(carrera,'Steeming_Bigram_Th_'+str(thresholds[thresholdIndex]),listTopics,dictionariesSteemingPerThreshold[thresholdIndex],corpusSteemingPerThreshold[thresholdIndex],datasetsSteemingBigram[thresholdIndex])
		coherencePerThresholdSteeming[thresholds[thresholdIndex]] = coherencePerKTopic

	writeExcelCoherencePerTopic(carrera,'Lemmatizacion_Bigram_Th',thresholds,listTopics,coherencePerThresholdLemma)
	writeExcelCoherencePerTopic(carrera,'Steeming_Bigram_Th',thresholds,listTopics,coherencePerThresholdSteeming)

def removeWordsFromDataset(dataset,words):
	newDataset = []
	for document in dataset:
		for wordToRemove in words:
			document = [wDoc for wDoc in document if wDoc!=wordToRemove]
		newDataset.append(document)
	return newDataset

def writeExcelCoherencePerTopic_TFIDF(career,typeProcessing,thresholds,porcDelete,numTopics,coherencePerThreshold):
	newExcel = openpyxl.Workbook()
	actualSheet = newExcel.active
	TypeProcessingColumn =1
	ThresholdColumn = 2
	PorcColumn = 3
	K_Column = 4
	CV_CoherenceColumn =5
	Umass_CoherenceColumn = 6
	CUCI_CoherenceColumn = 7
	CNPMI_CoherenceColumn = 8
	actualSheet.title = career 
	actualSheet.cell(row=1, column=TypeProcessingColumn).value = 'TypeProcessing'
	actualSheet.cell(row=1, column=ThresholdColumn).value = 'Threshold'
	actualSheet.cell(row=1, column=PorcColumn).value = 'PorcDelete'
	actualSheet.cell(row=1, column=K_Column).value = 'K_Topic'
	actualSheet.cell(row=1, column=CV_CoherenceColumn).value = 'c_v'
	actualSheet.cell(row=1, column=Umass_CoherenceColumn).value = 'u_mass'
	actualSheet.cell(row=1, column=CUCI_CoherenceColumn).value = 'c_uci'
	actualSheet.cell(row=1, column=CNPMI_CoherenceColumn).value = 'c_npmi'

	actualRow = 2
	for threshold in thresholds:
		for porc in porcDelete:
			for iTopic in numTopics:
				actualSheet.cell(row=actualRow, column=TypeProcessingColumn).value = typeProcessing
				actualSheet.cell(row=actualRow, column=ThresholdColumn).value = threshold
				actualSheet.cell(row=actualRow, column=PorcColumn).value = porc
				actualSheet.cell(row=actualRow, column=K_Column).value = iTopic
				actualSheet.cell(row=actualRow, column=CV_CoherenceColumn).value = coherencePerThreshold[threshold][porc][iTopic]['c_v']
				actualSheet.cell(row=actualRow, column=Umass_CoherenceColumn).value = coherencePerThreshold[threshold][porc][iTopic]['u_mass']
				actualSheet.cell(row=actualRow, column=CUCI_CoherenceColumn).value = coherencePerThreshold[threshold][porc][iTopic]['c_uci']
				actualSheet.cell(row=actualRow, column=CNPMI_CoherenceColumn).value = coherencePerThreshold[threshold][porc][iTopic]['c_npmi']

				actualRow += 1

	newExcel.save(career + '_' + typeProcessing + '_Resultados_Coherencia_TH_P_CV.xlsx')

def writeLDA_models_BIGRAM_TFIDF(career,typeProcessing,datasetsBigram,variances,thresholds,porcDelete,numTopics):
	coherencePerThreshold = {}

	for index_TH in range(len(thresholds)):
		print('Empezando con '+typeProcessing + ' '+ str(thresholds[index_TH]))
		wordsSortedByValue = sorted(variances[index_TH].items(),key=operator.itemgetter(1))
		wordPerPorc = []
		coherencePerThreshold[thresholds[index_TH]] = {}
		for porcIndex in range(len(porcDelete)):
			wordPerPorc.append([])
			coherencePerThreshold[thresholds[index_TH]][porcDelete[porcIndex]]={}

		bounds = []
		for porc in porcDelete:
			bounds.append(int(len(wordsSortedByValue)*porc/100))

		print('Extrayendo las palabras segun limites')
		for iWord in range(len(wordsSortedByValue)-1,len(wordsSortedByValue)-bounds[len(bounds)-1]-1,-1): # Se escoge el de mayor porcentaje			
			wordPerPorc[len(bounds)-1].append(wordsSortedByValue[iWord][0])
			for indexPorc in range(len(porcDelete)-1):
				if iWord >= len(wordsSortedByValue)-bounds[indexPorc]-1:
					wordPerPorc[indexPorc].append(wordsSortedByValue[iWord][0])

		for porcIndex in range(len(porcDelete)):
			print('Removiendo palabras TFIDF por porcentajes de ' + str(porcDelete[porcIndex]))
			dataset_TFIDF = removeWordsFromDataset(datasetsBigram[index_TH],wordPerPorc[porcIndex])
			pickle.dump(dataset_TFIDF,open("DatasetTFIDF_" + career + '_'+ typeProcessing + "_Th_" + str(thresholds[index_TH]) + '_Porc_' + str(porcDelete[porcIndex]) + ".p","wb"))

			dictionary = corpora.Dictionary(dataset_TFIDF)
			corpus = [dictionary.doc2bow(text) for text in dataset_TFIDF]

			print('Empenzando con los LDAs')
			for numTopic in numTopics:
				print('Obteniendo LDA de ' + typeProcessing + ' ' + str(porcDelete[porcIndex]) + ' NumTopic: ' + str(numTopic))
				ldamodel = gensim.models.ldamodel.LdaModel(corpus, num_topics=numTopic,id2word = dictionary, passes=20,random_state=42)
				print('Termino de construir LDA')
				#pickle.dump(ldamodel,open("LDA_TFIDF_" + career + '_'+ typeProcessing + "_Th_" + str(thresholds[index_TH]) + '_Porc_' + str(porcDelete[porcIndex]) + '_K_' + str(numTopic)+ ".p","wb"))
				F = open('LDA_TFIDF_'+career + '_'+ typeProcessing + "_Th_" + str(thresholds[index_TH]) + '_Porc_' + str(porcDelete[porcIndex]) + '_K_' + str(numTopic)+'.txt','w')
				cv_coherence_Model = CoherenceModel(model=ldamodel,corpus=corpus,texts=dataset_TFIDF,coherence='c_v').get_coherence()
				umass_coherence_Model = CoherenceModel(model=ldamodel,corpus=corpus,texts=dataset_TFIDF,coherence='u_mass').get_coherence()
				cuci_coherence_Model = CoherenceModel(model=ldamodel,corpus=corpus,texts=dataset_TFIDF,coherence='c_uci').get_coherence()
				cnpmi_coherence_Model = CoherenceModel(model=ldamodel,corpus=corpus,texts=dataset_TFIDF,coherence='c_npmi').get_coherence()
				

				coherencePerThreshold[thresholds[index_TH]][porcDelete[porcIndex]][numTopic]= {}				
				coherencePerThreshold[thresholds[index_TH]][porcDelete[porcIndex]][numTopic]['c_v'] = cv_coherence_Model
				coherencePerThreshold[thresholds[index_TH]][porcDelete[porcIndex]][numTopic]['u_mass'] = umass_coherence_Model
				coherencePerThreshold[thresholds[index_TH]][porcDelete[porcIndex]][numTopic]['c_uci'] = cuci_coherence_Model
				coherencePerThreshold[thresholds[index_TH]][porcDelete[porcIndex]][numTopic]['c_npmi'] = cnpmi_coherence_Model

				F.write('Coherencia c_v de ' + career + '_'+ typeProcessing + '_K_' + str(numTopic) + ': ' +str(cv_coherence_Model)+'\n')
				F.write('Coherencia u_mass de ' + career + '_'+ typeProcessing + '_K_' + str(numTopic) + ': ' +str(umass_coherence_Model)+'\n')
				F.write('Coherencia c_uci de ' + career + '_'+ typeProcessing + '_K_' + str(numTopic) + ': ' +str(cuci_coherence_Model)+'\n')
				F.write('Coherencia c_npmi de ' + career + '_'+ typeProcessing + '_K_' + str(numTopic) + ': ' +str(cnpmi_coherence_Model)+'\n')
				for iTopic in range(numTopic):
					F.write('Topic '+str(iTopic)+': '+str(ldamodel.print_topic(topicno=iTopic,topn=15))+'\n')
				F.close()

	print("Imprimiendo EXCEL de Coherencias")
	writeExcelCoherencePerTopic_TFIDF(career,typeProcessing,thresholds,porcDelete,numTopics,coherencePerThreshold)

def LDA_GENSIM_COMPLETO_BIGRAM_TFIDF():
	carrera = 'Informatica'

	datasetsLemmatizacionBigram = []
	datasetsStemmingBigram = []

	thresholds_Lemma = [15,24]
	thresholds_Stem = [18,24]

	variancesLemmatizacion = []
	variancesStemming = []

	numTopics = [7,8,9,10,11,12,13]
	porcDelete = [1,2,3,4,5,6,7]

	typeProcessing = 'Lemmatizacion'
	print('Cargando datasets y Variance de ' + typeProcessing)
	for threshold in thresholds_Lemma:
		datasetLemmatizacionBigram = pickle.load(open("DatasetBigram_Filter_DQ_V2_"+ carrera + "_" + typeProcessing +"_Th_"+ str(threshold)+".p","rb"))
		datasetsLemmatizacionBigram.append(datasetLemmatizacionBigram)
		varianceLemmatizacion = pickle.load(open("Variance_Filter_DQ_V2_"+ carrera + "_" + typeProcessing +"_Th_"+ str(threshold)+".p","rb"))
		variancesLemmatizacion.append(varianceLemmatizacion)

	writeLDA_models_BIGRAM_TFIDF(carrera,'Lemmatizacion_Filter_DQ_V2',datasetsLemmatizacionBigram,variancesLemmatizacion,thresholds_Lemma,porcDelete,numTopics)
	

	typeProcessing = 'Stemming'
	print('Cargando datasets y Variance de ' + typeProcessing)
	for threshold in thresholds_Stem:
		datasetStemmingBigram = pickle.load(open("DatasetBigram_Filter_DQ_"+ carrera + "_" + typeProcessing +"_Th_"+ str(threshold)+".p","rb"))
		datasetsStemmingBigram.append(datasetStemmingBigram)
		varianceStemming = pickle.load(open("Variance_Filter_DQ_"+ carrera + "_" + typeProcessing +"_Th_"+ str(threshold)+".p","rb"))
		variancesStemming.append(varianceStemming)

	

	writeLDA_models_BIGRAM_TFIDF(carrera,'Stemming_Filter_DQ_V2',datasetsStemmingBigram,variancesStemming,thresholds_Stem,porcDelete,numTopics)

def testLDA_randomState():
	career = 'Informatica'
	typeProcessing = 'Lemmatizacion'
	threshold = 15
	numTopic = 7
	porcDelete = 3
	datasetBigram = pickle.load(open("DatasetBigram_Filter_DQ_V2_"+ career + "_" + typeProcessing +"_Th_"+ str(threshold)+".p","rb"))
	varianceLemmatizacion = pickle.load(open("Variance_Filter_DQ_V2_"+ career + "_" + typeProcessing +"_Th_"+ str(threshold)+".p","rb"))
	wordsSortedByValue = sorted(varianceLemmatizacion.items(),key=operator.itemgetter(1))
	bound = int(len(wordsSortedByValue)*porcDelete/100)
	print('Extrayendo las palabras segun limites')
	wordPerPorc = []
	for iWord in range(len(wordsSortedByValue)-1,len(wordsSortedByValue)-bound-1,-1): # Se escoge el de mayor porcentaje			
		wordPerPorc.append(wordsSortedByValue[iWord][0])

	print('Removiendo palabras TFIDF por porcentajes de ' + str(porcDelete))
	dataset_TFIDF = removeWordsFromDataset(datasetBigram,wordPerPorc)
	
	dictionary = corpora.Dictionary(dataset_TFIDF)
	corpus = [dictionary.doc2bow(text) for text in dataset_TFIDF]

	print('Obteniendo LDA V1 de ' + typeProcessing + ' ' + str(porcDelete) + ' NumTopic: ' + str(numTopic))
	ldamodel = gensim.models.ldamodel.LdaModel(corpus, num_topics=numTopic,id2word = dictionary, passes=20,random_state=42)
	print('Termino de construir LDA')
	cv_coherence_Model = CoherenceModel(model=ldamodel,corpus=corpus,texts=dataset_TFIDF,coherence='c_v').get_coherence()
	umass_coherence_Model = CoherenceModel(model=ldamodel,corpus=corpus,texts=dataset_TFIDF,coherence='u_mass').get_coherence()
	cuci_coherence_Model = CoherenceModel(model=ldamodel,corpus=corpus,texts=dataset_TFIDF,coherence='c_uci').get_coherence()
	cnpmi_coherence_Model = CoherenceModel(model=ldamodel,corpus=corpus,texts=dataset_TFIDF,coherence='c_npmi').get_coherence()
		
	print('Coherencia c_v de ' + career + '_'+ typeProcessing + '_K_' + str(numTopic) + ': ' +str(cv_coherence_Model))
	print('Coherencia u_mass de ' + career + '_'+ typeProcessing + '_K_' + str(numTopic) + ': ' +str(umass_coherence_Model))
	print('Coherencia c_uci de ' + career + '_'+ typeProcessing + '_K_' + str(numTopic) + ': ' +str(cuci_coherence_Model))
	print('Coherencia c_npmi de ' + career + '_'+ typeProcessing + '_K_' + str(numTopic) + ': ' +str(cnpmi_coherence_Model))
	for iTopic in range(numTopic):
		print('Topic '+str(iTopic)+': '+str(ldamodel.print_topic(topicno=iTopic,topn=15)))

	print('Obteniendo LDA V2 de ' + typeProcessing + ' ' + str(porcDelete) + ' NumTopic: ' + str(numTopic))
	ldamodel2 = gensim.models.ldamodel.LdaModel(corpus, num_topics=numTopic,id2word = dictionary, passes=20,random_state=42)
	print('Termino de construir LDA')
	cv_coherence_Model = CoherenceModel(model=ldamodel2,corpus=corpus,texts=dataset_TFIDF,coherence='c_v').get_coherence()
	umass_coherence_Model = CoherenceModel(model=ldamodel2,corpus=corpus,texts=dataset_TFIDF,coherence='u_mass').get_coherence()
	cuci_coherence_Model = CoherenceModel(model=ldamodel2,corpus=corpus,texts=dataset_TFIDF,coherence='c_uci').get_coherence()
	cnpmi_coherence_Model = CoherenceModel(model=ldamodel2,corpus=corpus,texts=dataset_TFIDF,coherence='c_npmi').get_coherence()
		
	print('Coherencia c_v de ' + career + '_'+ typeProcessing + '_K_' + str(numTopic) + ': ' +str(cv_coherence_Model))
	print('Coherencia u_mass de ' + career + '_'+ typeProcessing + '_K_' + str(numTopic) + ': ' +str(umass_coherence_Model))
	print('Coherencia c_uci de ' + career + '_'+ typeProcessing + '_K_' + str(numTopic) + ': ' +str(cuci_coherence_Model))
	print('Coherencia c_npmi de ' + career + '_'+ typeProcessing + '_K_' + str(numTopic) + ': ' +str(cnpmi_coherence_Model))
	for iTopic in range(numTopic):
		print('Topic '+str(iTopic)+': '+str(ldamodel2.print_topic(topicno=iTopic,topn=15)))


#testLDA_randomState()
LDA_GENSIM_COMPLETO_BIGRAM_TFIDF()
#LDA_GENSIM_COMPLETO_BIGRAM()
#LDA_GENSIM_COMPLETO()
#LDA_GENSIM_DIVIDIDO()