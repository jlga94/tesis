from nltk.corpus import stopwords
import csv
import operator
import codecs
import string
import math
import unicodedata
import treetaggerwrapper
import openpyxl
import re
import numpy
import warnings
from nltk.stem.snowball import SpanishStemmer
from langdetect import detect
from nltk.tokenize import sent_tokenize

class TextProcessor:
    lemmatizer=None
    stopEnglish=None
    stopSpanish=None
    spanishStemmer=None

    def __init__(self):
        self.lemmatizer = treetaggerwrapper.TreeTagger(TAGLANG='es')
        self.stopEnglish = stopwords.words('english')
        self.stopSpanish = stopwords.words('spanish')
        self.stopSpanish.append('y/o')
        self.spanishStemmer=SpanishStemmer()
        #string.punctuation+='•'

    def _remove_numbers(self, text):
        "Elimina los números del texto"

        return ''.join([letter for letter in text if not letter.isdigit() and letter!='•'])

    def _remove_punctuation(self, text):
        "Elimina los signos de puntuacion del texto"

        regex = re.compile('[%s]' % re.escape(string.punctuation))
        return regex.sub(' ', text)

    def _separateFirstLetterFromPunctuation(self,text):
    	newText=''
    	previousLetter='1'
    	for letter in text:
    		if(previousLetter in string.punctuation and letter.isalpha()):
    			newText+=' '   			
    		newText+=letter
    		previousLetter=letter
    	return newText

    def _retireRareLetters(self,text):
    	return ''.join([letter for letter in text if letter.isalpha() or letter in string.punctuation or letter==' '])

    def _retireUrlsWWW(self,text):
    	return re.sub(r'\s*(?:https?://)?www\.\S*\.[A-Za-z]{2,5}\s*', '', text).strip()

    def _retireUrlsHTTP(self,text):
    	return re.sub(r'\w+:\/{2}[\d\w-]+(\.[\d\w-]+)*(?:(?:\/[^\s/]*))*', '', text).strip()

    def firstpreprocessText(self,text):
    	text=text.lower()
    	text=self._retireRareLetters(text)
    	text=self._retireUrlsWWW(text)
    	text=self._retireUrlsHTTP(text)
    	return text

    def preprocessText(self,text):
    	text=self._remove_punctuation(text)
    	text=self._remove_numbers(text)
    	return text

    def lematizeText(self,text):
    	newText = ""
    	lemmaElement = 2
    	
    	for sentence in sent_tokenize(text,language='spanish'): #Para 
    		textWithSeparatePunctuation=self._separateFirstLetterFromPunctuation(sentence)
    		lemmaResultText = self.lemmatizer.tag_text(textWithSeparatePunctuation)# Return [[word,type of word, lemma]]
    		#print(lemmaResultText)
    		for wordLemmaResult in lemmaResultText:
    			#print(word)
    			word = wordLemmaResult.split('\t')[lemmaElement]
    			if word not in self.stopEnglish and word not in self.stopSpanish:
    				newText += ' ' + word
    	
    	return newText.strip()

    def stemText(self,text):
        text=self._separateFirstLetterFromPunctuation(text)
        newText = ""
        firstWord = True
        for word in text.split():
            if word not in self.stopEnglish and word not in self.stopSpanish:
                word = word.replace("\ufeff", "")
                wordStemmed = self.spanishStemmer.stem(word)
                if firstWord:
                    newText += wordStemmed
                    firstWord = False
                else:
                    newText += " " + wordStemmed
        return newText

class Input_Output:
    procesadorTextos=None

    def __init__(self):
        self.procesadorTextos=TextProcessor()
        self.secciones={}
        self.nombColumnas=[]

    def limpiarDataset_Steeming(self, dataset):
    	newDataset = []
    	for ofertaText in dataset:
    		ofertaText = self.procesadorTextos.preprocessText(ofertaText)
    		ofertaText = self.procesadorTextos.stemText(ofertaText)
    		listaPalabras = ofertaText.strip().split()
    		newDataset.append(listaPalabras)
    	print('Se termino de limpiarlo completamente con Steeming.')
    	return newDataset

    def limpiarDataset_Lemmatizacion(self, dataset):
    	newDataset = []
    	for oferta in dataset:
    		cadenaLemmatizada = self.procesadorTextos.lematizeText(oferta.strip())
    		cadenaPreProcesadaLemmatizada = self.procesadorTextos.preprocessText(cadenaLemmatizada)
    		listaPalabras = cadenaPreProcesadaLemmatizada.split()
    		newDataset.append(listaPalabras)
    	print('Se termino de limpiarlo completamente con Lemmatizacion.')
    	return newDataset

    def leerSecciones(self, filename):
    	dataset = []
    	wb = openpyxl.load_workbook(filename)
    	sheets = wb.get_sheet_names()
    	sheetAviso = wb.get_sheet_by_name(sheets[0])
    	maxColumnas = sheetAviso.max_column + 1
    	primeraFila = 1
    	for iColumna in range(1,maxColumnas):
    		seccion = str(sheetAviso.cell(row=primeraFila, column=iColumna).value)
    		self.nombColumnas.append(seccion)
    		self.secciones[seccion]=iColumna
    
    def buscarNombre_Descripcion_Conocimientos_Ofertas():
    	columnasABuscar=[]
    	columnasABuscar.append(self.secciones['Job: Job Title'])
    	columnasABuscar.append(self.secciones['Job: Description'])
    	columnasABuscar.append(self.secciones['Job: Qualifications'])
    	return columnasABuscar

    def obtenerDatasetAClasificar_Completo(self, filename):
        "Se lee un archivo Excel con las ofertas a clasificar"

        dataset = []
        wb = openpyxl.load_workbook(filename)
        sheets = wb.get_sheet_names()
        sheetAviso = wb.get_sheet_by_name(sheets[0])
        maxFilas = sheetAviso.max_row + 1
        maxColumnas = sheetAviso.max_column + 1

        self.leerSecciones(filename)
        columnasABuscar=[]
        columnasABuscar.append(self.secciones['Job: Job Title'])
        columnasABuscar.append(self.secciones['Job: Description'])
        columnasABuscar.append(self.secciones['Job: Qualifications'])

        for num_oferta in range(2, maxFilas):
            fila = []
            completeText = ''
            for nColumna in columnasABuscar:
            	text = str(sheetAviso.cell(row=num_oferta, column=nColumna).value)
            	completeText+=' ' + text.lower()
            idioma = detect(completeText.strip())
            if(idioma=='es'):
            	text = self.procesadorTextos.firstpreprocessText(completeText)
            	dataset.append(text)
        print('Se termino de leer el Excel y pre procesarlo por primera vez.')
        #return self.limpiarDataset_Steeming(dataset)
        return self.limpiarDataset_Lemmatizacion(dataset)

    def obtenerDatasetAClasificar(self, filename):
        "Se lee un archivo Excel con las ofertas a clasificar"

        dataset = []
        wb = openpyxl.load_workbook(filename)
        sheets = wb.get_sheet_names()
        sheetAviso = wb.get_sheet_by_name(sheets[0])
        maxFilas = sheetAviso.max_row + 1

        for num_oferta in range(1, maxFilas):
            text = str(sheetAviso.cell(row=num_oferta, column=1).value)
            dataset.append(text)
        
        datasetLemmatizacion = self.limpiarDataset_Lemmatizacion(dataset)
        datasetSteeming = self.limpiarDataset_Steeming(dataset)

        return datasetLemmatizacion,datasetSteeming

    def contarIdiomasEnOfertas(self, filename):
    	dataset = []
    	wb = openpyxl.load_workbook(filename)
    	sheets = wb.get_sheet_names()
    	sheetAviso = wb.get_sheet_by_name(sheets[0])
    	maxFilas = sheetAviso.max_row + 1
    	maxColumnas = sheetAviso.max_column + 1

    	self.leerSecciones(filename)
    	columnasABuscar=[]
    	columnasABuscar.append(self.secciones['Job: Job Title'])
    	columnasABuscar.append(self.secciones['Job: Description'])
    	columnasABuscar.append(self.secciones['Job: Qualifications'])

    	ofertasRaras={}
    	idiomas={}
    	for num_oferta in range(2, maxFilas):
    		completeText = ''
    		for nColumna in columnasABuscar:
    			text = str(sheetAviso.cell(row=num_oferta, column=nColumna).value)
    			#completeText+=self.procesadorTextos.firstpreprocessText(text)
    			completeText+=' '+text.lower()
    		idioma = detect(completeText.strip())
    		if(idioma not in idiomas.keys()):
    			idiomas[idioma]=1
    		else:
    			idiomas[idioma]+=1
    		if(idioma!='es' and idioma!='en'):
    			if(idioma not in ofertasRaras.keys()):
    				ofertasRaras[idioma]=[completeText]
    			else:
    				ofertasRaras[idioma].append(completeText)
    	return idiomas,ofertasRaras
            


    def primeraLimpiezaADatasetAClasificar(self, filename):
    	wb = openpyxl.load_workbook(filename)
    	sheets = wb.get_sheet_names()
    	sheetAviso = wb.get_sheet_by_name(sheets[0])
    	maxFilas = sheetAviso.max_row + 1
    	maxColumnas = sheetAviso.max_column + 1

    	newExcel = openpyxl.Workbook()
    	newSheet = newExcel.active

    	self.leerSecciones(filename)
    	columnasABuscar=[]
    	columnasABuscar.append(self.secciones['Job: Job Title'])
    	columnasABuscar.append(self.secciones['Job: Description'])
    	columnasABuscar.append(self.secciones['Job: Qualifications'])

    	numColNewExcel=1
    	for num_oferta in range(2, maxFilas):
    		completeText = ''
    		for nColumna in columnasABuscar:
    			text = str(sheetAviso.cell(row=num_oferta, column=nColumna).value)
    			completeText+=' ' + text.lower()
    		idioma = detect(completeText.strip())
    		if(idioma=='es'):
    			text = self.procesadorTextos.firstpreprocessText(completeText)
    			newSheet.cell(row=numColNewExcel, column=1).value = text
    			numColNewExcel+=1

    	filenameWithoutExtension = filename.split('.')[0]
    	newExcel.save(filenameWithoutExtension + '_PrimerPreProcesamiento.xlsx')

